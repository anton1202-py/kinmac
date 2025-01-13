from datetime import datetime
import pandas as pd

import pandas as pd
from database.models import Articles, Marketplace, CostPrice, SalesReportOnSales
from django.db.models import Case, Count, IntegerField, Q, Sum, When
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side


def template_for_article_costprice(costprice_data):
    """Создает и скачивает excel файл с шаблоном для себестоимости"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active

    # Заполняем лист данными
    row = 2
    for article in costprice_data:
        ws.cell(row=row, column=1, value=article.article.common_article)
        ws.cell(row=row, column=2, value=article.article.name)
        if article.costprice:
            ws.cell(row=row, column=3, value=article.costprice)
        else:
            ws.cell(row=row, column=3, value="")
        row += 1
    # Устанавливаем заголовки столбцов
    ws.cell(row=1, column=1, value="Артикул")
    ws.cell(row=1, column=2, value="Название")
    ws.cell(row=1, column=3, value="Себестоимость")

    al = Alignment(horizontal="center", vertical="center")
    al_left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18

    for i in range(len(costprice_data) + 1):
        for c in ws[f"A{i+1}:C{i+1}"]:
            for i in range(3):
                c[i].border = Border(top=thin, left=thin, bottom=thin, right=thin)
                c[i].alignment = al_left

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type="application/xlsx")
    name = f'Article_Costprice_Template_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = "attachment; filename=" + name
    response["Content-Disposition"] = file_data
    wb.save(response)

    return response


def costprice_article_timport_from_excel(xlsx_file):
    """Импортирует данные о группе артикула из Excel"""
    excel_data_common = pd.read_excel(xlsx_file)
    column_list = excel_data_common.columns.tolist()
    if (
        "Артикул" in column_list
        and "Название" in column_list
        and "Себестоимость" in column_list
    ):
        excel_data = pd.DataFrame(
            excel_data_common, columns=["Артикул", "Название", "Себестоимость"]
        )
        article_list = excel_data["Артикул"].to_list()
        designer_type_list = excel_data["Название"].to_list()
        costprice_list = excel_data["Себестоимость"].to_list()

        for i in range(len(article_list)):

            if Articles.objects.filter(common_article=article_list[i]).exists():
                if CostPrice.objects.filter(
                    article__common_article=article_list[i]
                ).exists():
                    if not CostPrice.objects.filter(
                        article__common_article=article_list[i],
                        costprice=costprice_list[i],
                    ).exists():
                        CostPrice.objects.filter(
                            article__common_article=article_list[i]
                        ).update(
                            costprice=costprice_list[i], costprice_date=datetime.now()
                        )
                else:
                    CostPrice(
                        article__common_article=article_list[i],
                        costprice=costprice_list[i],
                        costprice_date=datetime.now(),
                    )
            else:
                return f"В базе данных нет артикула {article_list[i]}."

    else:
        return f"Вы пытались загрузить ошибочный файл {xlsx_file}."


def common_analytic_excel_file_export(data):
    """Создает и скачивает excel файл с общей аналитикой"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active
    table_headers = (
        "Артикул",
        "Средняя цена до СПП",
        "Реализация (сумма продаж до СПП)",
        "К перечислению",
        "Продажи",
        "Возвраты",
        "Себестоимость продаж",
        "Штрафы",
        "Компенсация подмененного",
        "Возмещение издержек по перевозке",
        "Оплата бракованного и потерянного",
        "Логистика",
        "Средняя стоимость логистики",
        "Хранение",
        "Кратность короба",
        "Услуга ФФ",
        "Рекламная кампания",
        "Самовыкуп",
        "Количество отказов и возвратов",
        "Продажи, шт",
        "Общее количество продаж с учетом возвратов",
        "Средний процент выкупа",
        "Средняя прибыль на 1 шт",
        "Налог",
        "Прибыль",
        "Прибыль с учетом самовыкупов",
        "ROI",
        "Рентабельность",
    )

    for col, header in enumerate(table_headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Заполняем лист данными
    for row, item in enumerate(data, start=2):
        ws.cell(row=row, column=1, value=item.article.common_article)
        ws.cell(row=row, column=2, value=item.average_price_before_spp)
        ws.cell(row=row, column=3, value=item.realization_summ_sale)
        ws.cell(row=row, column=4, value=item.for_pay)
        ws.cell(row=row, column=5, value=item.sale)
        ws.cell(row=row, column=6, value=item.returns)

        ws.cell(row=row, column=7, value=item.costprice_of_sales)
        ws.cell(row=row, column=8, value=item.penalty)
        ws.cell(row=row, column=9, value=item.compensation_for_the_substituted)
        ws.cell(row=row, column=10, value=item.reimbursement_of_transportation_costs)
        ws.cell(row=row, column=11, value=item.payment_defective_and_lost)

        ws.cell(row=row, column=12, value=item.logistic)
        ws.cell(row=row, column=13, value=item.average_logistic_cost)
        ws.cell(row=row, column=14, value=item.storage)
        ws.cell(row=row, column=15, value=item.box_multiplicity)
        ws.cell(row=row, column=16, value=item.ff_service)

        ws.cell(row=row, column=17, value=item.advertisment)
        ws.cell(row=row, column=18, value=item.self_purchase)
        ws.cell(row=row, column=19, value=item.refusals_and_returns_amount)
        ws.cell(row=row, column=20, value=item.sales_amount)
        ws.cell(row=row, column=21, value=item.common_sales_with_returns)

        ws.cell(row=row, column=22, value=item.average_percent_of_buyout)
        ws.cell(row=row, column=23, value=item.average_profit_for_one_piece)
        ws.cell(row=row, column=24, value=item.tax)
        ws.cell(row=row, column=25, value=item.profit)
        ws.cell(row=row, column=26, value=item.profit_with_self_purchase)

        ws.cell(row=row, column=27, value=item.roi)
        ws.cell(row=row, column=28, value=item.profitability)

    al = Alignment(horizontal="center", vertical="center")
    al_left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="000000")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 15
    ws.column_dimensions["J"].width = 15
    ws.column_dimensions["K"].width = 15
    ws.column_dimensions["L"].width = 15
    ws.column_dimensions["M"].width = 15
    ws.column_dimensions["N"].width = 15
    ws.column_dimensions["O"].width = 15
    ws.column_dimensions["P"].width = 15
    ws.column_dimensions["Q"].width = 15
    ws.column_dimensions["R"].width = 15
    ws.column_dimensions["S"].width = 15
    ws.column_dimensions["T"].width = 15
    ws.column_dimensions["U"].width = 15
    ws.column_dimensions["V"].width = 15
    ws.column_dimensions["W"].width = 15
    ws.column_dimensions["X"].width = 15
    ws.column_dimensions["Y"].width = 15
    ws.column_dimensions["Z"].width = 15
    ws.column_dimensions["AA"].width = 15
    ws.column_dimensions["AB"].width = 15

    for i in range(len(data) + 1):
        for c in ws[f"A{i+1}:AB{i+1}"]:
            for i in range(28):
                c[i].border = Border(top=thin, left=thin, bottom=thin, right=thin)
                c[i].alignment = al_left

    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type="application/xlsx")
    name = f'Common_analytic_{datetime.now().strftime("%Y.%m.%d")}.xlsx'
    file_data = "attachment; filename=" + name
    response["Content-Disposition"] = file_data
    wb.save(response)

    return response
