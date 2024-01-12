import datetime
import os
from datetime import date

import pandas as pd
import telegram
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.db.models import Q, Sum
from django.http import FileResponse, HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.views.generic import UpdateView
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from telegram_working.assistance import (save_message_function,
                                         upgrade_message_function)
from telegram_working.start_tg_approve import start_tg_working

from .forms import (ApprovalStatusForm, CashPaymentForm,
                    FilterPayWithCheckingForm, PaymentsForm, PayWithCardForm,
                    PayWithCheckingAccountForm, TransferToCardForm)
from .models import (ApprovalStatus, ApprovedFunction, CashPayment,
                     Contractors, PayerOrganization, Payments, PayWithCard,
                     PayWithCheckingAccount, TelegramApproveButtonMessage,
                     TelegramMessageActions, TransferToCard)
from .validators import StripToNumbers


def excel_creating_mod(data):
    """Создает и скачивает excel файл"""
    # Создаем DataFrame из данных
    wb = Workbook()
    # Получаем активный лист
    ws = wb.active

    # Заполняем лист данными
    for row, item in enumerate(data, start=2):
        print(item.pub_date)
        try:
            date_create = datetime.datetime.strptime(f'{item.pub_date}', '%Y-%m-%d %H:%M:%S.%f')
        except:
            date_create = datetime.datetime.strptime(f'{item.pub_date}', '%Y-%m-%d %H:%M:%S')
        ws.cell(row=row, column=1, value=date_create.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=2, value=item.payment_sum)
        if item.comment:
            ws.cell(row=row, column=3, value=str(item.comment))
        else:
            ws.cell(row=row, column=3, value='')
        if item.payer_organization:
            ws.cell(row=row, column=4, value=str(item.payer_organization.name))
        else:
             ws.cell(row=row, column=4, value='')
        ws.cell(row=row, column=5, value=str(item.contractor_name.name))
        ws.cell(row=row, column=6, value=str(item.project.name))
        ws.cell(row=row, column=7, value=str(item.category.name))
        ws.cell(row=row, column=8, value=str(item.payment_method.method_name))
        if item.date_of_payment:
            try:
                date_pay = datetime.datetime.strptime(f'{item.date_of_payment}', '%Y-%m-%d %H:%M:%S')
            except:
                date_pay = datetime.datetime.strptime(f'{item.date_of_payment}', '%Y-%m-%d %H:%M:%S.%f')
            ws.cell(row=row, column=9, value=date_pay.strftime('%d.%m.%Y'))
        else:
            ws.cell(row=row, column=9, value='')
    # Устанавливаем заголовки столбцов
    ws.cell(row=1, column=1, value='Дата')
    ws.cell(row=1, column=2, value='Сумма')
    ws.cell(row=1, column=3, value='Комментарий')
    ws.cell(row=1, column=4, value='Плательщик')
    ws.cell(row=1, column=5, value='Контрагент')
    ws.cell(row=1, column=6, value='Проект')
    ws.cell(row=1, column=7, value='Категория')
    ws.cell(row=1, column=8, value='Способ платежа')
    ws.cell(row=1, column=9, value='Оплачено')

    al = Alignment(horizontal="center",
        vertical="center")
    al_left = Alignment(horizontal="left",
        vertical="center")
    al2 = Alignment(vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    pattern = PatternFill('solid', fgColor="fcff52")

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 12

    for i in range(len(data)+1):
        for c in ws[f'A{i+1}:I{i+1}']:
            for i in range(9):
                c[i].border = Border(top=thin, left=thin,
                    bottom=thin, right=thin)
                c[i].alignment = al_left
            
    # Сохраняем книгу Excel в память
    response = HttpResponse(content_type='application/xlsx')
    response['Content-Disposition'] = 'attachment; filename=table.xlsx'
    wb.save(response)

    return response